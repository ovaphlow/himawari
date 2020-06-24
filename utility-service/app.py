from concurrent import futures
import json
import logging
import os
import uuid

from openpyxl import load_workbook
import grpc

import file_pb2
import file_pb2_grpc
import archive_pb2
import archive_pb2_grpc


class File(file_pb2_grpc.FileServicer):

    def ParseXlsx(self, request, context):
        print('准备导入文件', request.file_name)
        file_path = os.path.join('temp_file', request.file_name)
        wb = load_workbook(filename=file_path)
        for sheet in wb:
            print('数据行数（包括标题行）', len(tuple(sheet.rows)), '行')
            for i in range(2, len(tuple(sheet.rows)) + 1):
                # 判断关键数据是否有效
                if len(sheet.cell(rows=i, columns=4).value) != 18:
                    with grpc.insecure_channel('127.0.0.1:8912') as channel:
                        stub = message_pb2_grpc.MessageStub(channel)
                        response = stub.Save(message_pb2.SaveRequest(
                            uuid=str(uuid.uuid1()),
                            user_id=0,# request.user_id
                            doc=json.dumps({
                                'send_by': '导入档案服务',
                                'title': '身份证长度错误',
                                'content': '身份证 %s, 档案号 %s, 姓名 %s' % (
                                    sheet.cell(row=i, column=4).value,
                                    sheet.cell(row=i, column=2).value,
                                    sheet.cell(row=i, column=3).value
                                )
                            })
                        ))
                    continue
                if sheet.cell(row=i, column=2).value == '' or sheet.cell(row=i, column=3).value == '':
                    with grpc.insecure_channel('127.0.0.1:8912') as channel:
                        stub = message_pb2_grpc.MessageStub(channel)
                        response = stub.Save(message_pb2.SaveRequest(
                            uuid=str(uuid.uuid1()),
                            user_id=0,# request.user_id
                            doc=json.dumps({
                                'send_by': '导入档案服务',
                                'title': '档案号或姓名数据异常',
                                'content': '身份证 %s, 档案号 %s, 姓名 %s' % (
                                    sheet.cell(row=i, column=4).value,
                                    sheet.cell(row=i, column=2).value,
                                    sheet.cell(row=i, column=3).value
                                )
                            })
                        ))
                    continue
                # 检查档案号/身份证是否重复
                # 如有异常情况发送消息
                with grpc.insecure_channel('127.0.0.1:8911') as channel:
                    stub = archive_pb2_grpc.ArchiveStub(channel)
                    response = stub.Save(archive_pb2.SaveRequest(
                            uuid=str(uuid.uuid5(uuid.NAMESPACE_DNS, sheet.cell(row=i, column=4).value)),
                            sn=sheet.cell(row=i, column=2).value,
                            name=sheet.cell(row=i, column=3).value,
                            id_card=sheet.cell(row=i, column=4).value,
                            doc=json.dumps({
                                'bday': sheet.cell(row=i, column=5).value,
                                'tel': sheet.cell(row=i, column=6).value,
                                'remark': sheet.cell(row=i, column=7).value
                            })))

        resp = {}
        resp['message'] = ''
        resp['content'] = ''
        return file_pb2.Reply(data=json.dumps(resp))


def serve():
    server = grpc.server(futures.ThreadPoolExecutor(max_workers=8))
    file_pb2_grpc.add_FileServicer_to_server(File(), server)
    server.add_insecure_port('[::]:8913')
    server.start()
    print('UTILITY-SERVICE 运行于端口: %s' % (8913,))
    server.wait_for_termination()


if __name__ == '__main__':
    logging.basicConfig()
    serve()
